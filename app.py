import streamlit as st
import google.generativeai as genai
from PIL import Image
import json
import openpyxl
import io

# アプリのタイトル
st.title("📱 勤務表 自動入力アプリ")
st.write("iPhoneで撮った勤務表をアップロードすると、エクセルが完成します！")

# 1. 魔法の鍵を入れる場所
api_key = st.text_input("GeminiのAPIキー（AIza...）を入力してください", type="password")

# 2. 写真をアップロードする場所（複数OK）
uploaded_files = st.file_uploader("勤務表の写真（前半・後半）を選んでください", type=['jpg', 'jpeg', 'png'], accept_multiple_files=True)

# 3. 実行ボタン
if st.button("✨ エクセルを作成する"):
    if not api_key:
        st.error("APIキーを入力してください！")
    elif not uploaded_files:
        st.error("写真をアップロードしてください！")
    else:
        try:
            # AIの準備
            genai.configure(api_key=api_key)
            model = genai.GenerativeModel('gemini-2.5-flash')
            
            prompt = """
            この写真の勤務表から、日付、開始時間、終了時間を抜き出して。
            必ず以下のフォーマットで出力して。日付は必ず「月/日」の形式（例: 3/1、3/15）にして。
            [
              {"日付": "3/1", "開始": "09:00", "終了": "18:00"}
            ]
            """
            
            all_shifts = {}
            
            # ぐるぐる回る「処理中」のマークを出します
            with st.spinner("AIが写真を読み取っています... 数十秒お待ちください！"):
                for file in uploaded_files:
                    img = Image.open(file).convert('RGB')
                    response = model.generate_content([prompt, img])
                    
                    clean_text = response.text.replace('```json', '').replace('```', '').strip()
                    shift_data = json.loads(clean_text)
                    
                    for shift in shift_data:
                        date_str = shift.get("日付", "").strip()
                        parts = date_str.split('/')
                        if len(parts) == 2:
                            day = int(parts[1])
                            all_shifts[day] = shift
            
            # エクセルに書き込む
            if len(all_shifts) > 0:
                wb = openpyxl.load_workbook("template.xlsx")
                ws = wb.active
                
                for day, shift_info in all_shifts.items():
                    row = day + 5 # 1日は6行目
                    if 6 <= row <= 36:
                        ws.cell(row=row, column=5).value = shift_info.get("開始", "")
                        ws.cell(row=row, column=6).value = shift_info.get("終了", "")
                
                # エクセルデータをiPhoneに渡せる形に変換
                output = io.BytesIO()
                wb.save(output)
                output.seek(0)
                
                st.success("🎉 エクセルの作成が完了しました！下のボタンからダウンロードできます。")
                
                # ダウンロードボタンを表示
                st.download_button(
                    label="📥 エクセルをダウンロード",
                    data=output,
                    file_name="自動入力完了_勤務表.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                st.warning("写真からうまくデータを読み取れませんでした。")
                
        except Exception as e:
            st.error("エラーが発生しました。写真を変えてもう一度お試しください。")
